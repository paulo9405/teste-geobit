from django import forms
import datetime
from django_filters.rest_framework import DjangoFilterBackend
from core.models import Pessoa
from openpyxl import load_workbook
from rest_framework import viewsets
from .serializers import PessoaSerializer
from django.shortcuts import render, redirect


class MulherViewSet(viewsets.ModelViewSet):
    queryset = Pessoa.objects.filter(sexo='F', cidade='Meeren')
    serializer_class = PessoaSerializer
    http_method_names = ['get']


class PessoaViewSet(viewsets.ModelViewSet):
    queryset = Pessoa.objects.all().order_by('nascimento').reverse()
    serializer_class = PessoaSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ('sexo', )
    http_method_names = ['get']


class DocumentForm(forms.Form):
    docfile = forms.FileField(
        label='Select a file',
        help_text='max. 42 megabytes'
    )


def geobit(request):
    form = DocumentForm()
    return render(request, "core/form.html", {"form": form})


def upload_file(request):
    if request.method == "POST":
        form = DocumentForm(request.POST, request.FILES)

        if form.is_valid():
            filehandle = request.FILES['docfile']

            wb = load_workbook(filename=filehandle.file)
            for sheet in wb.worksheets:
                for val in sheet.iter_rows(
                        min_row=5,
                        max_row=21,
                        min_col=2,
                        max_col=12,
                        values_only=True):

                    print(str(sheet) + " " + str(val))

                    nome = val[1]
                    sobrenome = val[2]
                    sexo = val[3]
                    altura = val[4]
                    peso = val[5]

                    nascimento = datetime.datetime.fromtimestamp(val[6]).\
                        strftime('%Y-%m-%d')

                    bairro = val[7]
                    cidade = val[8]
                    estado = val[9]
                    numero = val[10]

                    Pessoa.objects.create(
                        nome=nome,
                        sobrenome=sobrenome,
                        sexo=sexo,
                        altura=altura,
                        peso=peso,
                        nascimento=nascimento,
                        bairro=bairro,
                        cidade=cidade,
                        estado=estado,
                        numero=numero
                    )
                if form.is_valid():
                    sucesso = 'Upload realizado com sucesso!'
                    form = DocumentForm()
                    data = {"form": form, 'sucesso': sucesso}
                    return render(request, "core/form.html", data)

            return redirect('core_geobit')
